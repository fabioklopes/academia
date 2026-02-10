import datetime
import json
from collections import defaultdict
from datetime import timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import login as auth_login, logout as auth_logout, authenticate
from django.core.exceptions import PermissionDenied
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, Case, When, Count
from .models import User, Turma, AttendanceRequest, TurmaAluno, Graduacao, PlanoAula, Pedido, Item, Meta as MetaModel, Log
from .forms import GraduacaoForm, ItemForm, PedidoForm, TurmaForm, SolicitacaoAcessoForm, PerfilEditForm, MetaForm
import calendar
import openpyxl
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa
from django.template.loader import get_template
from io import BytesIO
import uuid
import os
import time
from django.conf import settings
from .logs import create_log
from PIL import Image
from django.core.files.base import ContentFile
from django.views.decorators.http import require_POST
from django.contrib.auth.views import PasswordResetConfirmView
from django.urls import reverse_lazy

# --- HELPERS ---

BELT_IMAGES = {
    'WHITE': 'faixa_branca.png',
    'GRAY_WHITE': 'faixa_cinza_branca.png',
    'GRAY': 'faixa_cinza.png',
    'GRAY_BLACK': 'faixa_cinza_preta.png',
    'YELLOW_WHITE': 'faixa_amarela_branca.png',
    'YELLOW': 'faixa_amarela.png',
    'YELLOW_BLACK': 'faixa_amarela_preta.png',
    'ORANGE_WHITE': 'faixa_laranja_branca.png',
    'ORANGE': 'faixa_laranja.png',
    'ORANGE_BLACK': 'faixa_laranja_preta.png',
    'GREEN_WHITE': 'faixa_verde_branca.png',
    'GREEN': 'faixa_verde.png',
    'GREEN_BLACK': 'faixa_verde_preta.png',
    'BLUE': 'faixa_azul.png',
    'PURPLE': 'faixa_roxa.png',
    'BROWN': 'faixa_marrom.png',
    'BLACK': 'faixa_preta.png',
}

WEEKDAYS = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']

def get_student_stats(student, start_date, end_date):
    presencas_qs = AttendanceRequest.objects.filter(
        student=student,
        status='APR',
        attendance_date__range=[start_date, end_date]
    )
    
    total_presencas = 0
    for p in presencas_qs:
        if p.attendance_date.weekday() == 1:
            if p.class_type == 'BOTH':
                total_presencas += 2
            else:
                total_presencas += 1
        else:
            total_presencas += 1
    
    meta = MetaModel.objects.filter(
        data_inicio__lte=end_date,
        data_fim__gte=start_date
    ).first()
    
    total_aulas = 0
    total_ausencias = 0
    porcentagem = 0
    situacao = 'N/A'
    
    if meta:
        total_aulas = meta.meta_aulas
        if total_aulas > 0:
            porcentagem = (total_presencas / total_aulas) * 100
        
        total_ausencias = total_aulas - total_presencas
        if total_ausencias < 0: total_ausencias = 0
        
        if porcentagem < 20: situacao = 'Insatisfatório'
        elif porcentagem < 40: situacao = 'Regular'
        elif porcentagem < 60: situacao = 'Bom'
        elif porcentagem < 80: situacao = 'Satisfatório'
        else: situacao = 'Excelente'
        
    return {
        'total_aulas': total_aulas,
        'total_presencas': total_presencas,
        'total_ausencias': total_ausencias,
        'porcentagem': round(porcentagem, 1),
        'situacao': situacao
    }

def extract_class_type(reason):
    if '[TYPE: GI]' in reason: return 'Primeira Aula (Gi)'
    if '[TYPE: NOGI]' in reason: return 'Segunda Aula (No-Gi)'
    if '[TYPE: BOTH]' in reason: return 'As duas aulas'
    return 'N/A'

def get_class_description(class_type):
    if class_type == 'BOTH': return 'Integral'
    if class_type == 'GI': return 'Gi'
    if class_type == 'NOGI': return 'NoGi'
    return '-'

# --- VIEWS GERAIS ---

class CustomPasswordResetConfirmView(PasswordResetConfirmView):
    template_name = 'academia/password_reset_confirm.html'
    success_url = reverse_lazy('password_reset_complete')

    def form_valid(self, form):
        # The form's save method will hash the password and save the user.
        form.save()
        
        # Log the successful password reset.
        if hasattr(form, 'user') and form.user is not None:
            create_log(form.user, 'redefiniu a senha com sucesso')

        return super().form_valid(form)

def index(request):
    if request.user.is_authenticated:
        if request.user.is_student():
            return redirect('dashboard')
    
    context = {}
    if not request.user.is_authenticated:
        students = User.objects.filter(group_role='STD', status='ATIVO')
        professors = User.objects.filter(group_role='PRO', status='ATIVO')
        context.update({
            'students': students,
            'professors': professors,
        })
    else:
        context.update({
            'presencas_aprovadas': AttendanceRequest.objects.filter(student=request.user, status='APR').count(),
            'presencas_pendentes': AttendanceRequest.objects.filter(student=request.user, status='PEN').count(),
        })
        
    return render(request, 'academia/index.html', context)

def login_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        user = authenticate(request, username=email, password=password)
        if user is not None:
            if user.status == 'ATIVO':
                auth_login(request, user)
                create_log(user, 'realizou login')
                next_url = request.GET.get('next')
                return redirect(next_url) if next_url else redirect('dashboard')
            elif user.status == 'PENDENTE':
                create_log(user, 'tentou login (cadastro pendente)', status='FALHA')
                context = {'error': 'Seu cadastro está pendente de aprovação.'}
            else:
                create_log(user, 'tentou login (usuário inativo)', status='FALHA')
                context = {'error': 'Usuário inativo. Contate o administrador.'}
        else:
            create_log(None, 'tentativa de login falhou (credenciais inválidas)', status='FALHA', email=email)
            context = {'error': 'E-mail ou senha inválidos.'}
        return render(request, 'academia/login.html', context)
    return render(request, 'academia/login.html')

def logout_view(request):
    if request.user.is_authenticated:
        create_log(request.user, 'realizou logout')
    auth_logout(request)
    return redirect('index')

def solicitar_acesso(request):
    if request.method == 'POST':
        post_data = request.POST.copy()
        if post_data.get('has_responsible'):
            post_data['password'] = '12345678'
            post_data['password_confirm'] = '12345678'

        form = SolicitacaoAcessoForm(post_data, request.FILES)
        if form.is_valid():
            data = form.cleaned_data
            email = data['email']
            responsible_email = data.get('responsible_email')
            has_responsible = data.get('has_responsible')

            username = email
            if has_responsible and responsible_email:
                if responsible_email == email:
                    username = f"{email.split('@')[0]}+{uuid.uuid4().hex[:4]}@{email.split('@')[1]}"
                
                if User.objects.filter(username=username).exists():
                    messages.warning(request, 'Este e-mail já está sendo usado por outro usuário.')
                    return render(request, 'academia/solicitar_acesso.html', {'form': form})

            responsible_user = None
            if has_responsible and responsible_email:
                responsible_user = User.objects.filter(
                    username=responsible_email,
                    email=responsible_email,
                    group_role='STD',
                    status='ATIVO',
                    responsible__isnull=True
                ).first()
                if not responsible_user:
                    messages.error(request, f'O e-mail do responsável "{responsible_email}" não foi encontrado, não é de um aluno ativo, ou não é uma conta principal.')
                    return render(request, 'academia/solicitar_acesso.html', {'form': form})

            whatsapp_number = data.get('whatsapp')

            user = User(
                username=username, email=email,
                first_name=data['first_name'], last_name=data['last_name'],
                birthday=data['birthday'], whatsapp=whatsapp_number, status='PENDENTE',
                responsible=responsible_user,
                photo=data.get('photo')
            )
            
            user.set_password(data['password'])
            
            try:
                user.save()
                create_log(user, 'solicitou acesso ao sistema')
            except Exception as e:
                create_log(None, f'erro ao solicitar acesso: {str(e)}', status='FALHA', email=email, first_name=data['first_name'])
                return render(request, 'default_errors.html', {
                    'error_code': '500',
                    'error_message': f'Erro ao salvar usuário: {str(e)}'
                })
            
            messages.success(request, 'Sua solicitação de acesso foi enviada com sucesso! Aguarde a aprovação de um administrador.')
            return redirect('solicitar_acesso')
    else:
        form = SolicitacaoAcessoForm()

    return render(request, 'academia/solicitar_acesso.html', {'form': form})

@require_POST
def verificar_email_responsavel(request):
    try:
        data = json.loads(request.body)
        email_to_verify = data.get('email')
        if not email_to_verify:
            return JsonResponse({'isValid': False, 'message': 'O e-mail não foi fornecido.'}, status=400)

        responsible_user = User.objects.filter(
            username=email_to_verify,
            email=email_to_verify,
            status='ATIVO',
            group_role='STD',
            responsible__isnull=True
        ).first()

        if responsible_user:
            dependent_email_to_prefill = responsible_user.email
            return JsonResponse({
                'isValid': True,
                'message': f'Responsável encontrado: {responsible_user.get_full_name()}. O e-mail do dependente será preenchido.',
                'dependent_email': dependent_email_to_prefill
            })
        else:
            return JsonResponse({'isValid': False, 'message': 'Nenhum usuário ativo encontrado com este e-mail ou o usuário não é um responsável válido.'})
    except json.JSONDecodeError:
        return JsonResponse({'isValid': False, 'message': 'Requisição inválida.'}, status=400)
    except Exception as e:
        return JsonResponse({'isValid': False, 'message': f'Ocorreu um erro: {str(e)}'}, status=500)

@login_required
def switch_account(request, user_id):
    original_user_id = request.session.get('original_user_id', request.user.id)
    dependent_user = get_object_or_404(User, id=user_id, responsible_id=original_user_id)
    
    create_log(request.user, f'iniciou o gerenciamento da conta de {dependent_user.get_full_name()}')
    
    auth_login(request, dependent_user, backend='django.contrib.auth.backends.ModelBackend')
    
    request.session['original_user_id'] = original_user_id
    
    messages.info(request, f"Você agora está gerenciando a conta de {dependent_user.get_full_name()}.")
    return redirect('dashboard')

@login_required
def switch_account_back(request):
    original_user_id = request.session.get('original_user_id')
    if original_user_id:
        original_user = get_object_or_404(User, id=original_user_id)
        
        create_log(original_user, f'encerrou o gerenciamento da conta de {request.user.get_full_name()}')
        
        auth_login(request, original_user, backend='django.contrib.auth.backends.ModelBackend')
        messages.success(request, "Você voltou para a sua conta.")
    else:
        messages.warning(request, "Nenhuma conta original encontrada para retornar.")

    return redirect('dashboard')

@login_required
def dashboard(request):
    selected_month = request.GET.get('month')
    today = datetime.date.today()
    
    try:
        selected_month = int(selected_month) if selected_month else today.month
    except (ValueError, TypeError):
        selected_month = today.month

    aniversariantes = User.objects.filter(birthday__month=selected_month, status='ATIVO').order_by('birthday__day')
    
    month_names = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    months = [{'number': i + 1, 'name': name} for i, name in enumerate(month_names)]

    base_context = {
        'aniversariantes': aniversariantes,
        'selected_month': selected_month,
        'months': months,
    }

    if request.user.is_student():
        context = {
            'presencas_pendentes': AttendanceRequest.objects.filter(student=request.user, status='PEN').count(),
            'presencas_aprovadas': AttendanceRequest.objects.filter(student=request.user, status='APR').count(),
            'turmas': TurmaAluno.objects.filter(aluno=request.user),
        }
        
        meta_ativa = MetaModel.objects.filter(data_inicio__lte=today, data_fim__gte=today).first()
        if meta_ativa:
            presencas_no_periodo = AttendanceRequest.objects.filter(
                student=request.user,
                status='APR',
                attendance_date__range=[meta_ativa.data_inicio, meta_ativa.data_fim]
            ).count()
            
            percentual_presenca = (presencas_no_periodo / meta_ativa.meta_aulas) * 100 if meta_ativa.meta_aulas > 0 else 0
            
            if percentual_presenca < 20:
                classificacao = 'Insatisfatório'
                cor_classificacao = 'danger'
            elif percentual_presenca < 40:
                classificacao = 'Regular'
                cor_classificacao = 'warning'
            elif percentual_presenca < 60:
                classificacao = 'Bom'
                cor_classificacao = 'info'
            elif percentual_presenca < 80:
                classificacao = 'Satisfatório'
                cor_classificacao = 'primary'
            else:
                classificacao = 'Excelente'
                cor_classificacao = 'success'
                
            context.update({
                'meta_ativa': meta_ativa,
                'presencas_meta': presencas_no_periodo,
                'percentual_meta': round(percentual_presenca, 1),
                'classificacao_meta': classificacao,
                'cor_classificacao_meta': cor_classificacao,
            })

        context.update(base_context)
        return render(request, 'academia/dashboard_aluno.html', context)
    
    elif request.user.is_professor():
        context = {
            'turmas_count': Turma.objects.filter(ativa=True).count(),
            'solicitacoes_presenca': AttendanceRequest.objects.filter(status='PEN').count(),
            'alunos_ativos': User.objects.filter(group_role='STD', turmas__professor=request.user, status='ATIVO').distinct().count(),
            'alunos_pendentes': User.objects.filter(status='PENDENTE').count(),
            'pedidos_pendentes': Pedido.objects.filter(status='PEND').count(),
        }
        context.update(base_context)
        return render(request, 'academia/dashboard_professor.html', context)

    elif request.user.is_admin():
        context = {
            'total_students': User.objects.filter(group_role='STD', status='ATIVO').count(),
            'total_professors': User.objects.filter(group_role='PRO', status='ATIVO').count(),
            'total_active_turmas': Turma.objects.filter(ativa=True).count(),
            'pending_attendance_requests': AttendanceRequest.objects.filter(status='PEN').order_by('-attendance_date')[:5],
            'solicitacoes_presenca': AttendanceRequest.objects.filter(status='PEN').count(),
            'alunos_pendentes': User.objects.filter(status='PENDENTE').count(),
            'pedidos_pendentes': Pedido.objects.filter(status='PEND').count(),
        }
        context.update(base_context)
        return render(request, 'academia/dashboard_administrador.html', context)
        
    return redirect('index')

@login_required
def perfil(request):
    user_to_display = request.user
    original_user_id = request.session.get('original_user_id')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        user = user_to_display

        if action == 'update_kimono':
            height_str = request.POST.get('height')
            weight_str = request.POST.get('weight')

            user.height = int(height_str) if height_str else None
            user.weight = float(weight_str) if weight_str else None
            user.kimono_size = request.POST.get('kimono_size')
            user.belt_size = request.POST.get('belt_size')
            user.save()
            create_log(user, 'atualizou informações do kimono')
            messages.success(request, 'Informações do kimono atualizadas com sucesso!')
        
        elif action == 'change_password':
            current_password = request.POST.get('current_password')
            new_password = request.POST.get('new_password')
            confirm_new_password = request.POST.get('confirm_new_password')

            if not user.check_password(current_password):
                create_log(user, 'tentou alterar senha (senha atual incorreta)', status='FALHA')
                messages.error(request, 'Senha atual incorreta.')
            elif new_password != confirm_new_password:
                create_log(user, 'tentou alterar senha (senhas não coincidem)', status='FALHA')
                messages.error(request, 'As novas senhas não coincidem.')
            else:
                user.set_password(new_password)
                user.save()
                auth_login(request, user)
                create_log(user, 'alterou a senha')
                messages.success(request, 'Senha alterada com sucesso!')
        
        return redirect('perfil')

    context = {
        'graduacao': Graduacao.objects.filter(aluno=user_to_display).first(),
        'pedidos': Pedido.objects.filter(aluno=user_to_display).order_by('-data_solicitacao'),
        'user': user_to_display,
        'is_switched_account': original_user_id is not None,
        'original_user_id': original_user_id,
    }
    
    if original_user_id:
        original_user = User.objects.get(id=original_user_id)
        context['dependents'] = User.objects.filter(responsible=original_user)
        context['original_user'] = original_user

    return render(request, 'academia/perfil.html', context)

@login_required
def perfil_editar(request):
    if request.method == 'POST':
        form = PerfilEditForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            create_log(request.user, 'editou seu perfil')
            messages.success(request, 'Perfil atualizado com sucesso!')
            return redirect('perfil')
    else:
        form = PerfilEditForm(instance=request.user)
        
    return render(request, 'academia/perfil_editar.html', {'form': form})

@login_required
def perfil_photo_update(request):
    if request.method == 'POST':
        user = request.user
        new_photo = request.FILES.get('photo')

        if new_photo:
            if new_photo.size > 1048576:
                messages.error(request, 'A imagem selecionada é muito grande. O tamanho máximo é de 1MB.')
                return redirect('perfil')

            user.photo = new_photo
            user.save()
            create_log(user, 'atualizou a foto do perfil')
            messages.success(request, 'Foto do perfil atualizada com sucesso!')
        else:
            messages.error(request, 'Nenhuma foto foi selecionada.')

    return redirect('perfil')

# --- PAINEL DO ALUNO ---

@login_required
def aluno_marcar_presenca(request):
    if not request.user.is_student():
        raise PermissionDenied

    if request.method == 'POST':
        turma_id = request.POST.get('turma_id')
        dates = request.POST.getlist('dates')

        if not turma_id or not dates:
            messages.error(request, 'Turma e pelo menos uma data são obrigatórias.')
            return redirect('aluno_marcar_presenca')

        turma = get_object_or_404(Turma, id=turma_id)

        if not TurmaAluno.objects.filter(aluno=request.user, turma=turma).exists():
            messages.error(request, 'Você não está inscrito nesta turma.')
            return redirect('aluno_marcar_presenca')

        limit_date = timezone.localdate() - timedelta(days=15)

        for date_str in dates:
            try:
                attendance_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
                
                if attendance_date < limit_date:
                    messages.error(request, f'Não é possível solicitar presença para {attendance_date.strftime("%d/%m/%Y")}. O limite é de 15 dias retroativos.')
                    continue

                class_type_val = 'BOTH'
                if attendance_date.weekday() == 1: # Tuesday
                    class_type_val = request.POST.get(f'class_type_{date_str}', 'BOTH')
                
                reason_text = "Solicitação de presença pelo aluno."
                if class_type_val != 'BOTH':
                    reason_text += f" [TYPE: {class_type_val}]"
                elif attendance_date.weekday() == 1:
                    reason_text += " [TYPE: BOTH]"

                existing_request = AttendanceRequest.objects.filter(
                    student=request.user,
                    turma=turma,
                    attendance_date=attendance_date,
                    class_type=class_type_val
                ).first()

                if existing_request:
                    if existing_request.status == 'CAN':
                        existing_request.status = 'PEN'
                        existing_request.reason = reason_text
                        existing_request.class_type = class_type_val
                        existing_request.processed_by = None
                        existing_request.processed_at = None
                        existing_request.rejection_reason = ""
                        existing_request.notified = False
                        existing_request.save()
                        messages.success(request, f'Presença para {attendance_date.strftime("%d/%m/%Y")} solicitada com sucesso!')
                    else:
                        messages.warning(request, f'Já existe uma solicitação de presença ({existing_request.get_status_display()}) para {attendance_date.strftime("%d/%m/%Y")} nesta turma.')
                else:
                    AttendanceRequest.objects.create(
                        student=request.user, turma=turma,
                        attendance_date=attendance_date,
                        reason=reason_text,
                        class_type=class_type_val
                    )
                    messages.success(request, f'Presença para {attendance_date.strftime("%d/%m/%Y")} solicitada com sucesso!')
            except ValueError:
                messages.error(request, f'Formato de data inválido para {date_str}.')
                continue
        
        create_log(request.user, f'solicitou/alterou presença para {len(dates)} dia(s) na turma "{turma.nome}"')
        return redirect('aluno_presencas')

    today = timezone.localdate()
    year = today.year
    month = today.month
    limit_date = today - timedelta(days=15)

    all_attendance_requests = AttendanceRequest.objects.filter(student=request.user)
    
    attendance_data = {}
    for req in all_attendance_requests:
        date_str = req.attendance_date.strftime('%Y-%m-%d')
        if date_str not in attendance_data:
            attendance_data[date_str] = []
        attendance_data[date_str].append({
            'status': req.status,
            'turma': req.turma.nome,
            'reason': req.reason,
            'processed_by': req.processed_by.get_full_name() if req.processed_by else None,
            'processed_at': req.processed_at.strftime('%Y-%m-%d %H:%M') if req.processed_at else None,
            'rejection_reason': req.rejection_reason,
            'class_type': extract_class_type(req.reason),
        })

    context = {
        'turmas_aluno': TurmaAluno.objects.filter(aluno=request.user, status='APRO').select_related('turma'),
        'current_year': year,
        'current_month': month,
        'attendance_data_json': json.dumps(attendance_data),
        'limit_date': limit_date.strftime('%Y-%m-%d'),
        'today': today.strftime('%Y-%m-%d'),
    }
    return render(request, 'academia/aluno/marcar_presenca.html', context)

@login_required
def get_attendance_details(request):
    if not request.user.is_student():
        return JsonResponse({'error': 'Permissão negada'}, status=403)

    date_str = request.GET.get('date')
    if not date_str:
        return JsonResponse({'error': 'Data não fornecida'}, status=400)

    try:
        selected_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Formato de data inválido'}, status=400)

    attendance_requests = AttendanceRequest.objects.filter(
        student=request.user,
        attendance_date=selected_date
    ).select_related('turma', 'processed_by').order_by('turma__nome')

    details = []
    for req in attendance_requests:
        details.append({
            'turma': req.turma.nome,
            'status': req.get_status_display(),
            'reason': req.reason,
            'processed_by': req.processed_by.get_full_name() if req.processed_by else 'N/A',
            'processed_at': req.processed_at.strftime('%d/%m/%Y %H:%M') if req.processed_at else 'N/A',
            'rejection_reason': req.rejection_reason if req.rejection_reason else 'N/A',
            'class_type': extract_class_type(req.reason),
        })
    
    return JsonResponse({'date': date_str, 'details': details})

@login_required
def aluno_cancelar_presenca(request, request_id):
    if not request.user.is_student():
        raise PermissionDenied

    attendance_request = get_object_or_404(AttendanceRequest, id=request_id, student=request.user)

    if attendance_request.status != 'PEN':
        messages.error(request, 'Só é possível cancelar solicitações pendentes.')
    else:
        create_log(request.user, f'cancelou a solicitação de presença para {attendance_request.attendance_date.strftime("%d/%m/%Y")}')
        
        attendance_request.delete()
        
        messages.success(request, 'Solicitação de presença cancelada com sucesso.')

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'message': 'Cancelado com sucesso.'})

    return redirect('aluno_presencas')

@login_required
def aluno_presencas(request):
    if not request.user.is_student():
        raise PermissionDenied
    
    AttendanceRequest.objects.filter(student=request.user, status='APR', notified=False).update(notified=True)
    
    attendance_requests = AttendanceRequest.objects.filter(student=request.user).exclude(status='CAN').select_related('turma').order_by('-attendance_date')
    return render(request, 'academia/aluno/presencas.html', {'attendance_requests': attendance_requests})

@login_required
def aluno_graduacoes(request):
    if not request.user.is_student():
        raise PermissionDenied
    
    Graduacao.objects.filter(aluno=request.user, notified=False).update(notified=True)
    
    graduacoes = Graduacao.objects.filter(aluno=request.user).order_by('-data_graduacao')
    return render(request, 'academia/aluno/graduacoes.html', {'graduacoes': graduacoes})

@login_required
def aluno_pedidos(request):
    if not request.user.is_student():
        raise PermissionDenied
    pedidos = Pedido.objects.filter(aluno=request.user)
    return render(request, 'academia/aluno/pedidos.html', {'pedidos': pedidos})

@login_required
def aluno_pedido_novo(request):
    if not request.user.is_student():
        raise PermissionDenied
    
    if request.method == 'POST':
        form = PedidoForm(request.POST)
        if form.is_valid():
            pedido = form.save(commit=False)
            item = pedido.item
            
            pedido.aluno = request.user
            pedido.final_value = item.valor * pedido.quantidade if item.valor else None
            pedido.save()
            
            create_log(request.user, f'realizou pedido de {pedido.quantidade}x {item.nome}')
            messages.success(request, f'Pedido de "{item.nome}" realizado com sucesso! O item está reservado para você por 15 dias.')
            return redirect('aluno_pedidos')
    else:
        form = PedidoForm()

    itens = Item.objects.filter(quantidade__gt=0)
    item_data = {str(item.id): {'valor': str(item.valor), 'quantidade': item.quantidade} for item in itens}
    
    context = {'form': form, 'item_data_json': json.dumps(item_data)}
    return render(request, 'academia/aluno/pedido_form.html', context)

@login_required
def aluno_pedido_cancelar(request, pedido_id):
    if request.method == 'POST':
        pedido = get_object_or_404(Pedido, id=pedido_id, aluno=request.user)

        if pedido.status == 'PEND':
            item = pedido.item
            item.quantidade += pedido.quantidade
            item.save()

            pedido.status = 'CANC'
            pedido.save()
            create_log(request.user, f'cancelou pedido de {item.nome}')
            messages.success(request, 'Pedido cancelado com sucesso e item devolvido ao estoque.')
        else:
            messages.error(request, 'Apenas pedidos pendentes podem ser cancelados.')

    return redirect('aluno_pedidos')

@login_required
def aluno_relatorios(request):
    if not request.user.is_student():
        raise PermissionDenied
    return render(request, 'academia/aluno/relatorios.html')

@login_required
def gerar_relatorio_aluno(request):
    if not request.user.is_student():
        raise PermissionDenied

    report_type = request.GET.get('report_type')
    date_range = request.GET.get('date_range')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    today = datetime.date.today()
    start_date, end_date = None, None

    if date_range == 'current_month':
        start_date = today.replace(day=1)
        end_date = today.replace(day=calendar.monthrange(today.year, today.month)[1])
    elif date_range == 'custom':
        try:
            if start_date_str:
                start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            if end_date_str:
                end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Formato de data inválido. Use AAAA-MM-DD.")
            return redirect('perfil')

    context = {
        'report_type': report_type, 'start_date': start_date, 'end_date': end_date,
        'report_data': [], 'report_title': '',
        'no_data_message': 'Nenhum dado encontrado para o período selecionado.'
    }

    if report_type == 'frequencia':
        context['report_title'] = 'Relatório de Frequência de Aulas'
        query = AttendanceRequest.objects.filter(student=request.user, status='APR')
        if start_date: query = query.filter(attendance_date__gte=start_date)
        if end_date: query = query.filter(attendance_date__lte=end_date)
        context['report_data'] = query.order_by('attendance_date').select_related('turma')
        if not context['report_data'].exists():
            context['no_data_message'] = 'Nenhuma presença aprovada encontrada.'

    elif report_type == 'graduacao':
        context['report_title'] = 'Relatório de Progresso de Graduação'
        query = Graduacao.objects.filter(aluno=request.user)
        if start_date: query = query.filter(data_graduacao__gte=start_date)
        if end_date: query = query.filter(data_graduacao__lte=end_date)
        context['report_data'] = query.order_by('data_graduacao')
        if not context['report_data'].exists():
            context['no_data_message'] = 'Nenhum registro de graduação encontrado.'

    else:
        messages.error(request, "Tipo de relatório inválido.")
        return redirect('perfil')
    
    create_log(request.user, f'gerou relatório de {report_type}')
    return render(request, 'academia/aluno/relatorio_detalhe.html', context)

@login_required
def aluno_relatorio_presenca(request):
    if not request.user.is_student():
        raise PermissionDenied

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    export_format = request.GET.get('export')
    order = request.GET.get('order', 'desc')

    today = datetime.date.today()
    start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else today.replace(day=1)
    end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today

    presencas = AttendanceRequest.objects.filter(
        student=request.user,
        attendance_date__range=[start_date, end_date]
    ).exclude(status='CAN').order_by('attendance_date')

    grouped_data = {}
    for presenca in presencas:
        date_key = presenca.attendance_date
        if date_key not in grouped_data:
            grouped_data[date_key] = []
        grouped_data[date_key].append(presenca)

    report_data = []
    for date_key, presencas_list in grouped_data.items():
        is_tuesday = date_key.weekday() == 1
        
        status_final = 'APR'
        motivo_final = ""
        qty_final = 0
        
        apr_count = 0
        rej_count = 0
        types_apr = []
        
        for p in presencas_list:
            if p.status == 'APR':
                apr_count += 1
                types_apr.append(p.class_type)
            elif p.status == 'REJ':
                rej_count += 1
        
        if is_tuesday:
            if apr_count > 0:
                if 'BOTH' in types_apr or ('GI' in types_apr and 'NOGI' in types_apr):
                    qty_final = 2
                    motivo_final = "Integral"
                    status_final = 'APR'
                elif 'GI' in types_apr:
                    qty_final = 1
                    motivo_final = "Presente apenas na aula com Kimono"
                    status_final = 'APR'
                elif 'NOGI' in types_apr:
                    qty_final = 1
                    motivo_final = "Presente apenas na aula sem Kimono"
                    status_final = 'APR'
                else:
                    qty_final = 1
                    motivo_final = get_class_description(types_apr[0])
                    status_final = 'APR'
            else:
                qty_final = 0
                motivo_final = "Aluno ausente neste dia"
                if rej_count > 0:
                    first_rej = next((p for p in presencas_list if p.status == 'REJ'), None)
                    status_final = 'REJ'
                    if first_rej and first_rej.rejection_reason:
                         motivo_final += f" ({first_rej.rejection_reason})"
                else:
                     status_final = 'PEN'
        else:
            if apr_count > 0:
                qty_final = 1
                motivo_final = get_class_description(presencas_list[0].class_type)
                status_final = 'APR'
            else:
                qty_final = 0
                p = presencas_list[0]
                descricao_aula = get_class_description(p.class_type)
                if p.status == 'REJ':
                    status_final = 'REJ'
                    if p.rejection_reason:
                        motivo_final = f"{descricao_aula} - Ausente ({p.rejection_reason})"
                    else:
                        motivo_final = f"{descricao_aula} - Ausente"
                else:
                    status_final = p.status
                    motivo_final = descricao_aula

        report_data.append({
            'data': date_key,
            'status': dict(AttendanceRequest.STATUS_CHOICES).get(status_final, status_final),
            'motivo': motivo_final,
            'qty': qty_final
        })
        
    reverse = True if order == 'desc' else False
    report_data.sort(key=lambda x: x['data'], reverse=reverse)

    if export_format == 'pdf':
        create_log(request.user, 'exportou relatório de presenças (PDF)')
        
        graduacao = Graduacao.objects.filter(aluno=request.user).first()
        belt_image = BELT_IMAGES.get(graduacao.faixa, 'faixa_branca.png') if graduacao else 'faixa_branca.png'
        stats = get_student_stats(request.user, start_date, end_date)
        
        template = get_template('academia/aluno/relatorio_presenca_pdf.html')
        context = {
            'report_data': report_data,
            'report_title': 'Relatório de Presenças',
            'start_date': start_date,
            'end_date': end_date,
            'user': request.user,
            'graduacao': graduacao,
            'belt_image': belt_image,
            'stats': stats
        }
        html = template.render(context)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="relatorio_presencas.pdf"'
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('We had some errors <pre>' + html + '</pre>')
        return response

    if export_format == 'xlsx':
        create_log(request.user, 'exportou relatório de presenças (XLSX)')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="relatorio_presencas.xlsx"'
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Relatório de Presenças'
        
        headers = ["Data", "Status", "Motivo", "Quantidade"]
        for col_num, header_title in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header_title
            worksheet.column_dimensions[get_column_letter(col_num)].width = 20

        for row_num, data in enumerate(report_data, 2):
            weekday = WEEKDAYS[data['data'].weekday()]
            worksheet.cell(row=row_num, column=1, value=f"{data['data'].strftime('%d/%m/%Y')} {weekday}")
            worksheet.cell(row=row_num, column=2, value=data['status'])
            worksheet.cell(row=row_num, column=3, value=data['motivo'])
            worksheet.cell(row=row_num, column=4, value=data['qty'])

        workbook.save(response)
        return response

    items_per_page = request.GET.get('items_per_page', 10)
    try:
        items_per_page = int(items_per_page)
    except ValueError:
        items_per_page = 10

    paginator = Paginator(report_data, items_per_page)
    page = request.GET.get('page')
    try:
        report_data_page = paginator.page(page)
    except PageNotAnInteger:
        report_data_page = paginator.page(1)
    except EmptyPage:
        report_data_page = paginator.page(paginator.num_pages)

    context = {
        'report_data': report_data_page,
        'page_obj': report_data_page,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'items_per_page': items_per_page,
        'order': order,
    }
    return render(request, 'academia/aluno/relatorio_presenca.html', context)

@login_required
def aluno_relatorio_pedidos(request):
    if not request.user.is_student():
        raise PermissionDenied
    
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    export_format = request.GET.get('export')

    today = datetime.date.today()
    start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else today.replace(day=1)
    end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today

    pedidos = Pedido.objects.filter(
        aluno=request.user,
        data_solicitacao__date__range=[start_date, end_date]
    ).order_by('-data_solicitacao')

    if export_format == 'pdf':
        create_log(request.user, 'exportou relatório de pedidos (PDF)')
        template = get_template('academia/aluno/relatorio_pedidos_pdf.html')
        context = {
            'pedidos': pedidos,
            'report_title': 'Relatório de pedidos',
            'start_date': start_date,
            'end_date': end_date,
            'user': request.user
        }
        html = template.render(context)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="relatorio_pedidos.pdf"'
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('We had some errors <pre>' + html + '</pre>')
        return response

    if export_format == 'xlsx':
        create_log(request.user, 'exportou relatório de pedidos (XLSX)')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="relatorio_pedidos.xlsx"'
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Relatório de Pedidos'
        
        headers = ["Data do pedido", "Produto", "Quantidade", "Valor Total", "Status", "Desfecho"]
        for col_num, header_title in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header_title
            worksheet.column_dimensions[get_column_letter(col_num)].width = 25

        for row_num, pedido in enumerate(pedidos, 2):
            desfecho = pedido.rejection_reason or pedido.cancellation_reason or "-"
            worksheet.cell(row=row_num, column=1, value=pedido.data_solicitacao.strftime('%d/%m/%Y %H:%M'))
            worksheet.cell(row=row_num, column=2, value=pedido.aluno.get_full_name())
            worksheet.cell(row=row_num, column=3, value=pedido.item.nome)
            worksheet.cell(row=row_num, column=4, value=pedido.quantidade)
            worksheet.cell(row=row_num, column=5, value=pedido.final_value)
            worksheet.cell(row=row_num, column=6, value=pedido.get_status_display())
            worksheet.cell(row=row_num, column=7, value=desfecho)

        workbook.save(response)
        return response

    items_per_page = request.GET.get('items_per_page', 10)
    try:
        items_per_page = int(items_per_page)
    except ValueError:
        items_per_page = 10

    paginator = Paginator(pedidos, items_per_page)
    page = request.GET.get('page')
    try:
        pedidos_page = paginator.page(page)
    except PageNotAnInteger:
        pedidos_page = paginator.page(1)
    except EmptyPage:
        pedidos_page = paginator.page(paginator.num_pages)

    context = {
        'pedidos': pedidos_page,
        'page_obj': pedidos_page,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'items_per_page': items_per_page,
    }
    return render(request, 'academia/aluno/relatorio_pedidos.html', context)

# --- PAINEL DO PROFESSOR ---

@login_required
def professor_turmas(request):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    turmas = Turma.objects.all()
    return render(request, 'academia/professor/turmas.html', {'turmas': turmas})

@login_required
def professor_turma_nova(request):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    
    if request.method == 'POST':
        form = TurmaForm(request.POST)
        if form.is_valid():
            turma = form.save(commit=False)
            turma.professor = request.user
            turma.save()
            create_log(request.user, f'criou a turma "{turma.nome}"')
            messages.success(request, f'Turma "{turma.nome}" criada com sucesso!')
            return redirect('professor_turmas')
    else:
        form = TurmaForm()
    
    return render(request, 'academia/professor/turma_form.html', {'form': form})

@login_required
def professor_turma_editar(request, turma_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    
    turma = get_object_or_404(Turma, pk=turma_id)

    if request.method == 'POST':
        form = TurmaForm(request.POST, instance=turma)
        if form.is_valid():
            form.save()
            create_log(request.user, f'editou a turma "{turma.nome}"')
            messages.success(request, f'Turma "{turma.nome}" atualizada com sucesso!')
            return redirect('professor_turmas')
    else:
        form = TurmaForm(instance=turma)
    
    return render(request, 'academia/professor/turma_form.html', {'form': form, 'turma': turma})

@login_required
def professor_turma_alunos(request, turma_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    
    turma = get_object_or_404(Turma, pk=turma_id)
    
    alunos_na_turma_list = TurmaAluno.objects.filter(turma=turma, status='APRO').select_related('aluno').order_by('aluno__first_name', 'aluno__last_name')
    
    items_per_page = request.GET.get('items_per_page', 10)
    try:
        items_per_page = int(items_per_page)
    except ValueError:
        items_per_page = 10

    paginator = Paginator(alunos_na_turma_list, items_per_page)
    page = request.GET.get('page')
    
    try:
        alunos_turma = paginator.page(page)
    except PageNotAnInteger:
        alunos_turma = paginator.page(1)
    except EmptyPage:
        alunos_turma = paginator.page(paginator.num_pages)
        
    alunos_na_turma_ids = TurmaAluno.objects.filter(turma=turma, status='APRO').values_list('aluno_id', flat=True)
    
    context = {
        'turma': turma,
        'alunos_turma': alunos_turma,
        'alunos_disponiveis': User.objects.filter(group_role='STD', status='ATIVO').exclude(id__in=alunos_na_turma_ids).order_by('first_name', 'last_name'),
        'items_per_page': items_per_page,
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'academia/professor/partials/turma_alunos_list.html', context)

    return render(request, 'academia/professor/turma_alunos.html', context)

@login_required
def professor_turma_adicionar_aluno(request, turma_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    
    turma = get_object_or_404(Turma, pk=turma_id)
    
    if request.method == 'POST':
        alunos_ids = request.POST.getlist('alunos')
        if not alunos_ids:
            messages.error(request, 'Selecione pelo menos um aluno.')
        else:
            count = 0
            for aluno_id in alunos_ids:
                aluno = get_object_or_404(User, pk=aluno_id, group_role='STD')
                if not TurmaAluno.objects.filter(turma=turma, aluno=aluno).exists():
                    TurmaAluno.objects.create(turma=turma, aluno=aluno, status='APRO', data_aprovacao=timezone.now())
                    count += 1
            
            if count > 0:
                create_log(request.user, f'adicionou {count} aluno(s) à turma "{turma.nome}"')
                messages.success(request, f'{count} aluno(s) adicionado(s) à turma "{turma.nome}".')
            else:
                messages.warning(request, 'Nenhum aluno novo adicionado.')
        
    return redirect('professor_turma_alunos', turma_id=turma_id)

@login_required
def professor_turma_remover_aluno(request, turma_id, aluno_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    
    turma = get_object_or_404(Turma, pk=turma_id)
    
    turma_aluno = get_object_or_404(TurmaAluno, turma=turma, aluno_id=aluno_id)
    aluno_nome = turma_aluno.aluno.get_full_name()
    turma_aluno.delete()
    
    create_log(request.user, f'removeu o aluno "{aluno_nome}" da turma "{turma.nome}"')
    messages.success(request, f'Aluno "{aluno_nome}" removido da turma "{turma.nome}".')
    return redirect('professor_turma_alunos', turma_id=turma_id)

@login_required
def professor_alunos(request):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    
    status_order = Case(
        When(status='PENDENTE', then=1),
        When(status='ATIVO', then=2),
        When(status='INATIVO', then=3),
        default=4
    )
    
    alunos_list = User.objects.filter(group_role='STD').order_by(status_order, 'first_name')
    
    query = request.GET.get('q')
    if query:
        alunos_list = alunos_list.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        )
    
    items_per_page = request.GET.get('items_per_page', 10)
    try:
        items_per_page = int(items_per_page)
    except ValueError:
        items_per_page = 10

    paginator = Paginator(alunos_list, items_per_page)
    page = request.GET.get('page')
    
    try:
        alunos = paginator.page(page)
    except PageNotAnInteger:
        alunos = paginator.page(1)
    except EmptyPage:
        alunos = paginator.page(paginator.num_pages)
        
    context = {'alunos': alunos, 'query': query, 'items_per_page': items_per_page}
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'academia/professor/partials/aluno_list.html', context)
        
    return render(request, 'academia/professor/alunos.html', context)

@login_required
def promover_aluno(request):
    if not request.user.is_admin():
        raise PermissionDenied

    status_order = Case(
        When(status='PENDENTE', then=1),
        When(status='ATIVO', then=2),
        When(status='INATIVO', then=3),
        default=4
    )
    
    alunos_list = User.objects.filter(group_role='STD').order_by(status_order, 'first_name')
    
    query = request.GET.get('q')
    if query:
        alunos_list = alunos_list.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        )
    
    items_per_page = request.GET.get('items_per_page', 10)
    try:
        items_per_page = int(items_per_page)
    except ValueError:
        items_per_page = 10

    paginator = Paginator(alunos_list, items_per_page)
    page = request.GET.get('page')
    
    try:
        alunos = paginator.page(page)
    except PageNotAnInteger:
        alunos = paginator.page(1)
    except EmptyPage:
        alunos = paginator.page(paginator.num_pages)
        
    context = {'alunos': alunos, 'query': query, 'items_per_page': items_per_page}
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'academia/professor/partials/promover_aluno_list.html', context)
        
    return render(request, 'academia/professor/promover_aluno.html', context)

@login_required
def professor_aluno_desativar(request, aluno_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    aluno = get_object_or_404(User, pk=aluno_id, group_role='STD')
    aluno.status = 'INATIVO'
    aluno.save()
    create_log(request.user, f'desativou o usuário {aluno.get_full_name()}')
    messages.success(request, f'O usuário {aluno.get_full_name()} foi desativado.')
    return redirect('professor_alunos')

@login_required
def professor_aluno_excluir(request, aluno_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    aluno = get_object_or_404(User, pk=aluno_id)
    if aluno.status == 'PENDENTE':
        aluno.delete()
        create_log(request.user, f'excluiu a solicitação de cadastro de {aluno.get_full_name()}')
        messages.success(request, f'A solicitação de {aluno.get_full_name()} foi excluída.')
    else:
        messages.error(request, 'Apenas solicitações pendentes podem ser excluídas.')
    return redirect('professor_alunos')

@login_required
def professor_aluno_ativar(request, aluno_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    aluno = get_object_or_404(User, pk=aluno_id)
    previous_status = aluno.status
    aluno.status = 'ATIVO'
    aluno.save()
    if previous_status == 'PENDENTE':
        action_text = f'aprovou o cadastro de {aluno.get_full_name()}'
        messages.success(request, f'O cadastro de {aluno.get_full_name()} foi aprovado.')
    else:
        action_text = f'ativou o usuário {aluno.get_full_name()}'
        messages.success(request, f'O usuário {aluno.get_full_name()} foi ativado.')
    create_log(request.user, action_text)
    return redirect('professor_alunos')

@login_required
def professor_aluno_definir_tipo(request, aluno_id):
    if not request.user.is_admin():
        raise PermissionDenied

    user_to_change = get_object_or_404(User, pk=aluno_id)

    if request.method == 'POST':
        password = request.POST.get('password')
        if not request.user.check_password(password):
            create_log(request.user, f'tentou promover {user_to_change.get_full_name()} (senha incorreta)', status='FALHA')
            messages.error(request, 'Senha incorreta.')
            return redirect('promover_aluno')

        user_to_change.group_role = 'PRO'
        user_to_change.save()
        create_log(request.user, f'promoveu {user_to_change.get_full_name()} para Professor')
        messages.success(request, f'O tipo de usuário de {user_to_change.get_full_name()} foi alterado para Professor.')
        return redirect('promover_aluno')

    return redirect('promover_aluno')

@login_required
def tamanhos_medidas(request):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied

    students = User.objects.filter(group_role='STD').order_by('first_name', 'last_name')
    today = datetime.date.today()
    
    items_per_page = request.GET.get('items_per_page', 10)
    try:
        items_per_page = int(items_per_page)
    except ValueError:
        items_per_page = 10

    paginator = Paginator(students, items_per_page)
    page = request.GET.get('page')
    
    try:
        students_page = paginator.page(page)
    except PageNotAnInteger:
        students_page = paginator.page(1)
    except EmptyPage:
        students_page = paginator.page(paginator.num_pages)

    student_data = []
    for student in students_page:
        age = None
        if student.birthday:
            age = today.year - student.birthday.year - ((today.month, today.day) < (student.birthday.month, student.birthday.day))
        
        student_data.append({
            'user': student,
            'age': age
        })

    return render(request, 'academia/professor/tamanhos_medidas.html', {
        'student_data': student_data,
        'page_obj': students_page,
        'items_per_page': items_per_page
    })

@login_required
def professor_presencas(request):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied

    query = AttendanceRequest.objects.filter(status='PEN')
    
    pending_requests = query.select_related('student', 'turma').order_by('-attendance_date')
    return render(request, 'academia/professor/presencas.html', {'pending_requests': pending_requests})

@login_required
def professor_presenca_aprovar(request, request_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied

    attendance_request = get_object_or_404(AttendanceRequest, id=request_id)
    
    if request.method == 'POST':
        if not attendance_request.student.status == 'ATIVO':
            messages.error(request, f'Não é possível aprovar a presença de um aluno desativado ({attendance_request.student.get_full_name()}).')
            return redirect('professor_presencas')

        attendance_request.status = 'APR'
        attendance_request.processed_by = request.user
        attendance_request.processed_at = timezone.now()
        attendance_request.notified = False
        attendance_request.save()

        create_log(request.user, f'aprovou a presença de {attendance_request.student.get_full_name()} para a data {attendance_request.attendance_date.strftime("%d/%m/%Y")}')
        messages.success(request, 'Solicitação de presença aprovada.')
    
    return redirect('professor_presencas')

@login_required
def professor_presenca_rejeitar(request, request_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied

    attendance_request = get_object_or_404(AttendanceRequest, id=request_id)

    if request.method == 'POST':
        reason = request.POST.get('rejection_reason', 'Sem motivo especificado.')
        rejection_scope = request.POST.get('rejection_scope', 'both')
        
        is_tuesday = attendance_request.attendance_date.weekday() == 1
        is_both_classes = attendance_request.class_type == 'BOTH'
        
        if is_tuesday and is_both_classes:
            if rejection_scope == 'first':
                AttendanceRequest.objects.create(
                    student=attendance_request.student,
                    turma=attendance_request.turma,
                    attendance_date=attendance_request.attendance_date,
                    reason="Solicitação de presença pelo aluno. [TYPE: GI]",
                    class_type='GI',
                    status='REJ',
                    rejection_reason=reason,
                    processed_by=request.user,
                    processed_at=timezone.now()
                )
                
                attendance_request.reason = "Solicitação de presença pelo aluno. [TYPE: NOGI]"
                attendance_request.class_type = 'NOGI'
                attendance_request.status = 'APR'
                attendance_request.processed_by = request.user
                attendance_request.processed_at = timezone.now()
                attendance_request.notified = False
                attendance_request.save()
                
                create_log(request.user, f'rejeitou a 1ª aula e aprovou a 2ª aula de {attendance_request.student.get_full_name()} para a data {attendance_request.attendance_date.strftime("%d/%m/%Y")}')
                messages.success(request, 'Solicitação processada: 1ª aula rejeitada, 2ª aula aprovada.')
                return redirect('professor_presencas')
                
            elif rejection_scope == 'second':
                AttendanceRequest.objects.create(
                    student=attendance_request.student,
                    turma=attendance_request.turma,
                    attendance_date=attendance_request.attendance_date,
                    reason="Solicitação de presença pelo aluno. [TYPE: NOGI]",
                    class_type='NOGI',
                    status='REJ',
                    rejection_reason=reason,
                    processed_by=request.user,
                    processed_at=timezone.now()
                )
                
                attendance_request.reason = "Solicitação de presença pelo aluno. [TYPE: GI]"
                attendance_request.class_type = 'GI'
                attendance_request.status = 'APR'
                attendance_request.processed_by = request.user
                attendance_request.processed_at = timezone.now()
                attendance_request.notified = False
                attendance_request.save()
                
                create_log(request.user, f'rejeitou a 2ª aula e aprovou a 1ª aula de {attendance_request.student.get_full_name()} para a data {attendance_request.attendance_date.strftime("%d/%m/%Y")}')
                messages.success(request, 'Solicitação processada: 2ª aula rejeitada, 1ª aula aprovada.')
                return redirect('professor_presencas')

        attendance_request.status = 'REJ'
        attendance_request.rejection_reason = reason
        attendance_request.processed_by = request.user
        attendance_request.processed_at = timezone.now()
        attendance_request.save()
        
        create_log(request.user, f'rejeitou a presença de {attendance_request.student.get_full_name()} para a data {attendance_request.attendance_date.strftime("%d/%m/%Y")}')
        messages.success(request, 'Solicitação de presença rejeitada.')
    
    return redirect('professor_presencas')

@login_required
def professor_graduacoes(request):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied

    if request.user.is_admin():
        alunos_list = User.objects.filter(group_role='STD')
    else:
        alunos_list = User.objects.filter(group_role='STD', turmas__professor=request.user).distinct()
    
    items_per_page = request.GET.get('items_per_page', 10)
    try:
        items_per_page = int(items_per_page)
    except ValueError:
        items_per_page = 10

    paginator = Paginator(alunos_list, items_per_page)
    page = request.GET.get('page')
    
    try:
        alunos = paginator.page(page)
    except PageNotAnInteger:
        alunos = paginator.page(1)
    except EmptyPage:
        alunos = paginator.page(paginator.num_pages)

    graduacoes = Graduacao.objects.filter(aluno__in=alunos).distinct()
    
    context = {
        'alunos': alunos,
        'graduacoes': {grad.aluno.id: grad for grad in graduacoes},
        'items_per_page': items_per_page
    }
    return render(request, 'academia/professor/graduacoes.html', context)

@login_required
def professor_graduacao_editar(request, aluno_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
        
    aluno = get_object_or_404(User, id=aluno_id)
    graduacao, created = Graduacao.objects.get_or_create(aluno=aluno)

    if request.method == 'POST':
        form = GraduacaoForm(request.POST, instance=graduacao)
        if form.is_valid():
            graduacao = form.save(commit=False)
            graduacao.notified = False
            graduacao.save()
            create_log(request.user, f'atualizou a graduação de {aluno.get_full_name()}')
            messages.success(request, 'Graduação salva com sucesso!')
            return redirect('professor_graduacoes')
    else:
        form = GraduacaoForm(instance=graduacao)

    return render(request, 'academia/professor/graduacao_form.html', {'form': form, 'aluno': aluno})

@login_required
def professor_planos_aula(request):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    planos_aula = PlanoAula.objects.all() if request.user.is_admin() else PlanoAula.objects.filter(professor=request.user)
    return render(request, 'academia/professor/planos_aula.html', {'planos_aula': planos_aula})

@login_required
def professor_plano_aula_novo(request):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    return render(request, 'academia/professor/plano_aula_form.html')

@login_required
def professor_plano_aula_editar(request, plano_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    plano_aula = get_object_or_404(PlanoAula, pk=plano_id)
    if not request.user.is_admin() and plano_aula.professor != request.user:
        raise PermissionDenied
    return render(request, 'academia/professor/plano_aula_form.html', {'plano_aula': plano_aula})

@login_required
def professor_rankings(request):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    return render(request, 'academia/professor/rankings.html', {'rankings': []})

@login_required
def professor_ranking_novo(request):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    return render(request, 'academia/professor/ranking_form.html')

@login_required
def professor_pedidos(request):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    
    pedidos_list = Pedido.objects.all().order_by('-data_solicitacao')

    query = request.GET.get('q')
    if query:
        pedidos_list = pedidos_list.filter(
            Q(aluno__first_name__icontains=query) |
            Q(aluno__last_name__icontains=query) |
            Q(item__nome__icontains=query)
        )

    items_per_page = request.GET.get('items_per_page', 10)
    try:
        items_per_page = int(items_per_page)
    except ValueError:
        items_per_page = 10

    paginator = Paginator(pedidos_list, items_per_page)
    page = request.GET.get('page')

    try:
        pedidos = paginator.page(page)
    except PageNotAnInteger:
        pedidos = paginator.page(1)
    except EmptyPage:
        pedidos = paginator.page(paginator.num_pages)

    context = {'pedidos': pedidos, 'query': query, 'items_per_page': items_per_page}

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'academia/professor/partials/pedidos_list.html', context)

    return render(request, 'academia/professor/pedidos.html', context)

@login_required
def professor_pedido_aprovar(request, pedido_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    
    pedido = get_object_or_404(Pedido, pk=pedido_id)
    if request.method == 'POST':
        pedido.status = 'APRO'
        pedido.aprovado_por = request.user
        pedido.data_aprovacao = timezone.now()
        pedido.save()
        create_log(request.user, f'aprovou o pedido de {pedido.aluno.get_full_name()} ({pedido.item.nome})')
        messages.success(request, 'Pedido atendido com sucesso!')
    
    return redirect('professor_pedidos')

@login_required
def professor_pedido_rejeitar(request, pedido_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    
    pedido = get_object_or_404(Pedido, pk=pedido_id)
    if request.method == 'POST':
        if pedido.status == 'PEND':
            item = pedido.item
            item.quantidade += pedido.quantidade
            item.save()

        pedido.status = 'REJE'
        pedido.rejection_reason = request.POST.get('rejection_reason', 'Sem motivo especificado.')
        pedido.save()
        create_log(request.user, f'rejeitou o pedido de {pedido.aluno.get_full_name()} ({pedido.item.nome})')
        messages.success(request, 'Pedido rejeitado e estoque atualizado.')

    return redirect('professor_pedidos')

@login_required
def professor_pedido_cancelar(request, pedido_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    
    pedido = get_object_or_404(Pedido, pk=pedido_id)
    if request.method == 'POST':
        pedido.status = 'CANC'
        pedido.cancellation_reason = request.POST.get('cancellation_reason', 'Sem motivo especificado.')
        pedido.save()
        create_log(request.user, f'cancelou o pedido de {pedido.aluno.get_full_name()} ({pedido.item.nome})')
        messages.success(request, 'Pedido cancelado com sucesso!')

    return redirect('professor_pedidos')

@login_required
def professor_pedido_entregar(request, pedido_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    
    pedido = get_object_or_404(Pedido, pk=pedido_id)
    if request.method == 'POST':
        if pedido.item.quantidade is not None:
            if pedido.quantidade > pedido.item.quantidade:
                messages.error(request, 'Não há estoque suficiente para entregar este pedido.')
                return redirect('professor_pedidos')
            pedido.item.quantidade -= pedido.quantidade
            pedido.item.save()
        
        if final_value := request.POST.get('final_value'):
            pedido.final_value = final_value
        
        pedido.status = 'ENTR'
        pedido.save()
        create_log(request.user, f'marcou como entregue o pedido de {pedido.aluno.get_full_name()} ({pedido.item.nome})')
        messages.success(request, 'Pedido marcado como entregue!')

    return redirect('professor_pedidos')

@login_required
def professor_pedido_finalizar(request, pedido_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    
    pedido = get_object_or_404(Pedido, pk=pedido_id)
    if request.method == 'POST':
        pedido.status = 'FINA'
        pedido.save()
        create_log(request.user, f'finalizou o pedido de {pedido.aluno.get_full_name()} ({pedido.item.nome})')
        messages.success(request, 'Pedido finalizado com sucesso!')

    return redirect('professor_pedidos')

@login_required
def professor_itens(request):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    
    itens_list = Item.objects.all().order_by('nome')

    query = request.GET.get('q')
    if query:
        itens_list = itens_list.filter(
            Q(nome__icontains=query)
        )
    
    items_per_page = request.GET.get('items_per_page', 10)
    try:
        items_per_page = int(items_per_page)
    except ValueError:
        items_per_page = 10

    paginator = Paginator(itens_list, items_per_page)
    page = request.GET.get('page')
    
    try:
        itens = paginator.page(page)
    except PageNotAnInteger:
        itens = paginator.page(1)
    except EmptyPage:
        itens = paginator.page(paginator.num_pages)
        
    context = {'itens': itens, 'query': query, 'items_per_page': items_per_page}
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'academia/professor/partials/itens_list.html', context)
        
    return render(request, 'academia/professor/itens.html', context)

@login_required
def professor_item_novo(request):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    if request.method == 'POST':
        form = ItemForm(request.POST)
        if form.is_valid():
            item = form.save()
            create_log(request.user, f'criou o item "{item.nome}"')
            messages.success(request, 'Item criado com sucesso!')
            return redirect('professor_itens')
    else:
        form = ItemForm()
    return render(request, 'academia/professor/item_form.html', {'form': form})

@login_required
def professor_item_editar(request, item_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    item = get_object_or_404(Item, pk=item_id)
    if request.method == 'POST':
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            create_log(request.user, f'editou o item "{item.nome}"')
            messages.success(request, 'Item atualizado com sucesso!')
            return redirect('professor_itens')
    else:
        form = ItemForm(instance=item)
    return render(request, 'academia/professor/item_form.html', {'form': form})

@login_required
def professor_item_deletar(request, item_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    item = get_object_or_404(Item, pk=item_id)
    if request.method == 'POST':
        item_nome = item.nome
        item.delete()
        create_log(request.user, f'deletou o item "{item_nome}"')
        messages.success(request, 'Item deletado com sucesso!')
        return redirect('professor_itens')
    return render(request, 'academia/professor/item_confirm_delete.html', {'item': item})

@login_required
def professor_relatorios(request):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    return render(request, 'academia/professor/relatorios.html')

@login_required
def relatorio_pedidos(request):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied

    base_query = Pedido.objects.all()

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    group_by_aluno = request.GET.get('group_by_aluno')
    export_format = request.GET.get('export')

    today = datetime.date.today()
    start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else today.replace(day=1)
    end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today

    pedidos = base_query.filter(
        data_solicitacao__date__range=[start_date, end_date]
    ).order_by('data_solicitacao', 'aluno__first_name', 'aluno__last_name').select_related('aluno', 'item')

    if export_format == 'pdf':
        create_log(request.user, 'exportou relatório de pedidos (PDF)')
        template = get_template('academia/professor/relatorio_pedidos_pdf.html')
        context = {
            'pedidos': pedidos,
            'report_title': 'Relatório de Pedidos',
            'start_date': start_date,
            'end_date': end_date,
            'grouped': False
        }
        html = template.render(context)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="relatorio_pedidos.pdf"'
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('We had some errors <pre>' + html + '</pre>')
        return response

    if export_format == 'xlsx':
        create_log(request.user, 'exportou relatório de pedidos (XLSX)')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="relatorio_pedidos.xlsx"'
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Relatório de Pedidos'
        
        headers = ["Data do pedido", "Aluno", "Produto", "Quantidade", "Valor Total", "Status", "Desfecho"]
        for col_num, header_title in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header_title
            worksheet.column_dimensions[get_column_letter(col_num)].width = 25

        for row_num, pedido in enumerate(pedidos, 2):
            desfecho = pedido.rejection_reason or pedido.cancellation_reason or "-"
            worksheet.cell(row=row_num, column=1, value=pedido.data_solicitacao.strftime('%d/%m/%Y %H:%M'))
            worksheet.cell(row=row_num, column=2, value=pedido.aluno.get_full_name())
            worksheet.cell(row=row_num, column=3, value=pedido.item.nome)
            worksheet.cell(row=row_num, column=4, value=pedido.quantidade)
            worksheet.cell(row=row_num, column=5, value=pedido.final_value)
            worksheet.cell(row=row_num, column=6, value=pedido.get_status_display())
            worksheet.cell(row=row_num, column=7, value=desfecho)

        workbook.save(response)
        return response

    pedidos_data = pedidos
    page_obj = None

    items_per_page = request.GET.get('items_per_page', 10)
    try:
        items_per_page = int(items_per_page)
    except ValueError:
        items_per_page = 10

    if group_by_aluno:
        pedidos_agrupados = defaultdict(list)
        for pedido in pedidos:
            pedidos_agrupados[pedido.aluno.get_full_name()].append(pedido)
        
        pedidos_list = list(dict(pedidos_agrupados).items())
        
        paginator = Paginator(pedidos_list, items_per_page)
        page = request.GET.get('page')
        try:
            pedidos_page = paginator.page(page)
        except PageNotAnInteger:
            pedidos_page = paginator.page(1)
        except EmptyPage:
            pedidos_page = paginator.page(paginator.num_pages)
            
        pedidos_data = dict(pedidos_page.object_list)
        page_obj = pedidos_page
    else:
        paginator = Paginator(pedidos_data, items_per_page)
        page = request.GET.get('page')
        try:
            pedidos_page = paginator.page(page)
        except PageNotAnInteger:
            pedidos_page = paginator.page(1)
        except EmptyPage:
            pedidos_page = paginator.page(paginator.num_pages)
        
        pedidos_data = pedidos_page
        page_obj = pedidos_page

    context = {
        'pedidos': pedidos_data,
        'page_obj': page_obj,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'group_by_aluno': group_by_aluno,
        'alunos': User.objects.filter(group_role='STD'),
        'itens': Item.objects.all(),
        'items_per_page': items_per_page,
    }
    return render(request, 'academia/professor/relatorio_pedidos.html', context)

@login_required
def relatorio_presenca(request):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied

    turmas = Turma.objects.all()
    alunos = User.objects.filter(group_role='STD', status='ATIVO').order_by('first_name', 'last_name')
    base_query = AttendanceRequest.objects.all()

    turma_id = request.GET.get('turma')
    aluno_id = request.GET.get('aluno')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    export_format = request.GET.get('export')
    order = request.GET.get('order', 'desc')

    today = datetime.date.today()
    start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else today.replace(day=1)
    end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today

    query = base_query.filter(attendance_date__range=[start_date, end_date]).exclude(status='CAN')

    if turma_id and turma_id.isdigit():
        query = query.filter(turma_id=turma_id)
        turma_id = int(turma_id)
    else:
        turma_id = None

    if aluno_id and aluno_id.isdigit():
        query = query.filter(student_id=aluno_id)
        aluno_id = int(aluno_id)
    else:
        aluno_id = None

    presencas = query.order_by('attendance_date', 'student__first_name', 'student__last_name').select_related('student')

    grouped_data = {}
    for presenca in presencas:
        key = (presenca.attendance_date, presenca.student)
        if key not in grouped_data:
            grouped_data[key] = []
        grouped_data[key].append(presenca)

    report_data = []
    for (date_key, student), presencas_list in grouped_data.items():
        is_tuesday = date_key.weekday() == 1
        
        status_final = 'APR'
        motivo_final = ""
        qty_final = 0
        
        apr_count = 0
        rej_count = 0
        types_apr = []
        
        for p in presencas_list:
            if p.status == 'APR':
                apr_count += 1
                types_apr.append(p.class_type)
            elif p.status == 'REJ':
                rej_count += 1
        
        if is_tuesday:
            if apr_count > 0:
                if 'BOTH' in types_apr or ('GI' in types_apr and 'NOGI' in types_apr):
                    qty_final = 2
                    motivo_final = "Integral"
                    status_final = 'APR'
                elif 'GI' in types_apr:
                    qty_final = 1
                    motivo_final = "Presente apenas na aula com Kimono"
                    status_final = 'APR'
                elif 'NOGI' in types_apr:
                    qty_final = 1
                    motivo_final = "Presente apenas na aula sem Kimono"
                    status_final = 'APR'
                else:
                    qty_final = 1
                    motivo_final = get_class_description(types_apr[0])
                    status_final = 'APR'
            else:
                qty_final = 0
                motivo_final = "Aluno ausente neste dia"
                if rej_count > 0:
                    first_rej = next((p for p in presencas_list if p.status == 'REJ'), None)
                    status_final = 'REJ'
                    if first_rej and first_rej.rejection_reason:
                         motivo_final += f" ({first_rej.rejection_reason})"
                else:
                     status_final = 'PEN'
        else:
            if apr_count > 0:
                qty_final = 1
                motivo_final = get_class_description(presencas_list[0].class_type)
                status_final = 'APR'
            else:
                qty_final = 0
                p = presencas_list[0]
                descricao_aula = get_class_description(p.class_type)
                if p.status == 'REJ':
                    status_final = 'REJ'
                    if p.rejection_reason:
                        motivo_final = f"{descricao_aula} - Ausente ({p.rejection_reason})"
                    else:
                        motivo_final = f"{descricao_aula} - Ausente"
                else:
                    status_final = p.status
                    motivo_final = descricao_aula

        report_data.append({
            'data': date_key,
            'aluno': student.get_full_name(),
            'student_obj': student,
            'status': dict(AttendanceRequest.STATUS_CHOICES).get(status_final, status_final),
            'motivo': motivo_final,
            'qty': qty_final
        })
        
    reverse = True if order == 'desc' else False
    report_data.sort(key=lambda x: (x['data'], x['aluno']), reverse=reverse)

    if export_format == 'pdf':
        create_log(request.user, 'exportou relatório de presenças (PDF)')
        
        students_data = {}
        
        if aluno_id:
            student = get_object_or_404(User, id=aluno_id)
            students_data[student.id] = {
                'student': student,
                'presencas': [],
                'graduacao': Graduacao.objects.filter(aluno=student).first(),
                'stats': get_student_stats(student, start_date, end_date)
            }
            grad = students_data[student.id]['graduacao']
            students_data[student.id]['belt_image'] = BELT_IMAGES.get(grad.faixa, 'faixa_branca.png') if grad else 'faixa_branca.png'

        for item in report_data:
            s_id = item['student_obj'].id
            if s_id not in students_data:
                grad = Graduacao.objects.filter(aluno=item['student_obj']).first()
                students_data[s_id] = {
                    'student': item['student_obj'],
                    'presencas': [],
                    'graduacao': grad,
                    'belt_image': BELT_IMAGES.get(grad.faixa, 'faixa_branca.png') if grad else 'faixa_branca.png',
                    'stats': get_student_stats(item['student_obj'], start_date, end_date)
                }
            
            students_data[s_id]['presencas'].append({
                'data': item['data'],
                'status': item['status'],
                'motivo': item['motivo'],
                'qty': item['qty']
            })

        template = get_template('academia/professor/relatorio_presenca_pdf.html')
        context = {
            'students_data': list(students_data.values()),
            'report_title': 'Relatório de Presenças',
            'start_date': start_date,
            'end_date': end_date
        }
        html = template.render(context)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="relatorio_presencas.pdf"'
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('We had some errors <pre>' + html + '</pre>')
        return response

    if export_format == 'xlsx':
        create_log(request.user, 'exportou relatório de presenças (XLSX)')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="relatorio_presencas.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Relatório de Presenças'

        headers = ["Data", "Aluno", "Status", "Motivo", "Quantidade"]
        for col_num, header_title in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header_title
            worksheet.column_dimensions[get_column_letter(col_num)].width = 25

        for row_num, data in enumerate(report_data, 2):
            weekday = WEEKDAYS[data['data'].weekday()]
            worksheet.cell(row=row_num, column=1, value=f"{data['data'].strftime('%d/%m/%Y')} {weekday}")
            worksheet.cell(row=row_num, column=2, value=data['aluno'])
            worksheet.cell(row=row_num, column=3, value=data['status'])
            worksheet.cell(row=row_num, column=4, value=data['motivo'])
            worksheet.cell(row=row_num, column=5, value=data['qty'])

        workbook.save(response)
        return response

    items_per_page = request.GET.get('items_per_page', 10)
    try:
        items_per_page = int(items_per_page)
    except ValueError:
        items_per_page = 10

    paginator = Paginator(report_data, items_per_page)
    page = request.GET.get('page')
    try:
        report_data_page = paginator.page(page)
    except PageNotAnInteger:
        report_data_page = paginator.page(1)
    except EmptyPage:
        report_data_page = paginator.page(paginator.num_pages)

    context = {
        'turmas': turmas,
        'alunos': alunos,
        'report_data': report_data_page,
        'page_obj': report_data_page,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'turma_id': turma_id,
        'aluno_id': aluno_id,
        'items_per_page': items_per_page,
        'order': order,
    }
    return render(request, 'academia/professor/relatorio_presenca.html', context)

# --- VIEWS DE METAS ---

@login_required
def professor_metas(request):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    metas = MetaModel.objects.all()
    return render(request, 'academia/professor/metas.html', {'metas': metas})

@login_required
def professor_meta_nova(request):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    
    if request.method == 'POST':
        form = MetaForm(request.POST)
        if form.is_valid():
            meta = form.save(commit=False)
            meta.professor = request.user
            meta.save()
            create_log(request.user, f'criou a meta "{meta.titulo}"')
            messages.success(request, 'Meta criada com sucesso!')
            return redirect('professor_metas')
    else:
        form = MetaForm()
    
    return render(request, 'academia/professor/meta_form.html', {'form': form})

@login_required
def professor_meta_editar(request, meta_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    
    meta = get_object_or_404(MetaModel, pk=meta_id)
    
    if request.method == 'POST':
        form = MetaForm(request.POST, instance=meta)
        if form.is_valid():
            form.save()
            create_log(request.user, f'editou a meta "{meta.titulo}"')
            messages.success(request, 'Meta atualizada com sucesso!')
            return redirect('professor_metas')
    else:
        form = MetaForm(instance=meta)
    
    return render(request, 'academia/professor/meta_form.html', {'form': form})

@login_required
def professor_meta_deletar(request, meta_id):
    if not request.user.is_professor_or_admin():
        raise PermissionDenied
    
    meta = get_object_or_404(MetaModel, pk=meta_id)
    
    if request.method == 'POST':
        meta_titulo = meta.titulo
        meta.delete()
        create_log(request.user, f'excluiu a meta "{meta_titulo}"')
        messages.success(request, 'Meta excluída com sucesso!')
        return redirect('professor_metas')
    
    return render(request, 'academia/professor/meta_confirm_delete.html', {'meta': meta})

# --- LOGS ---

@login_required
def log_list(request):
    if not request.user.is_admin():
        raise PermissionDenied
    
    items_per_page = request.GET.get('items_per_page', 10)
    try:
        items_per_page = int(items_per_page)
        if items_per_page not in [10, 20, 30, 50]:
            items_per_page = 10
    except ValueError:
        items_per_page = 10

    status_filter = request.GET.get('status')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    logs_list = Log.objects.all().order_by('-timestamp')
    
    if status_filter:
        logs_list = logs_list.filter(status=status_filter)
    
    if start_date_str:
        try:
            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d')
            logs_list = logs_list.filter(timestamp__gte=start_date)
        except ValueError:
            pass

    if end_date_str:
        try:
            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
            logs_list = logs_list.filter(timestamp__lte=end_date)
        except ValueError:
            pass
    
    paginator = Paginator(logs_list, items_per_page)
    page = request.GET.get('page')
    
    try:
        logs = paginator.page(page)
    except PageNotAnInteger:
        logs = paginator.page(1)
    except EmptyPage:
        logs = paginator.page(paginator.num_pages)
        
    return render(request, 'academia/logs.html', {
        'logs': logs,
        'items_per_page': items_per_page,
        'status_filter': status_filter,
        'start_date': start_date_str,
        'end_date': end_date_str
    })

# --- VIEWS DE ERRO ---

def error_404(request, exception):
    return render(request, 'default_errors.html', {
        'error_code': '404',
        'error_message': 'O endereço pode estar incorreto ou a página foi removida.'
    }, status=404)

def error_500(request):
    return render(request, 'default_errors.html', {
        'error_code': '500',
        'error_message': 'Erro interno do servidor. Tente novamente mais tarde.'
    }, status=500)

def error_403(request, exception):
    return render(request, 'default_errors.html', {
        'error_code': '403',
        'error_message': 'Você não tem permissão para acessar esta página.'
    }, status=403)

def error_400(request, exception):
    return render(request, 'default_errors.html', {
        'error_code': '400',
        'error_message': 'Requisição inválida.'
    }, status=400)

def error_413(request, exception):
    return render(request, 'default_errors.html', {
        'error_code': '413',
        'error_message': 'Arquivo muito grande ou dados excedendo limites.'
    }, status=413)
